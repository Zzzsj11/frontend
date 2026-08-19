from __future__ import annotations

import json
import uuid
from datetime import timedelta
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, Field
from sqlalchemy import case, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from .auth import CurrentUser
from .config import settings
from .database import database_session
from .general_outline_comparison import compare_general_outlines
from .jobs import jobs as job_manager
from .kling import ASPECT_RATIOS as KLING_ASPECT_RATIOS
from .kling import IMAGE_TYPES as KLING_IMAGE_TYPES
from .kling import MAX_DURATION as KLING_MAX_DURATION
from .kling import MIN_DURATION as KLING_MIN_DURATION
from .kling import MODES as KLING_MODES
from .kling import KlingError
from .kling import create_task as kling_create_task
from .kling import query_task as kling_query_task
from .llm_comparison import CHAT_TEST_MODELS, compare_chat_models
from .models import (
    AdminOperationLogModel,
    AiModelModel,
    AiProviderModel,
    ApiErrorLogModel,
    ApiRequestLogModel,
    DigitalHumanModel,
    GenerationJobModel,
    H3TestPresetModel,
    LlmCallLogModel,
    ProjectCastModel,
    ProjectModel,
    ProjectTaskModel,
    PromptTemplateModel,
    PromptVersionModel,
    ServerMaintenanceRunModel,
    ShotAssetModel,
    SongEmotionProfileModel,
    StoryboardLineModel,
    StoryboardOptionItemModel,
    TokenUsageModel,
    UserModel,
    utcnow,
)
from .prompts import DEFAULT_PROMPTS, invalidate, render_lenient, template_variables
from .providers import ProviderError, list_video_models, query_provider_task, resume_generation, store_provider_result
from .rbac import (
    SONG_EMOTIONS_MANAGE,
    SONG_EMOTIONS_READ,
    STORYBOARD_OPTIONS_MANAGE,
    STORYBOARD_OPTIONS_READ,
    require_permission,
    require_super_admin,
)
from .runninghub import (
    ASPECT_RATIOS,
    DEFAULT_FIRST_FRAME_MEGAPIXELS,
    DEFAULT_STAGE1_MEGAPIXELS,
    DEFAULT_STAGE2_MEGAPIXELS,
    DEFAULT_TEXT_MEGAPIXELS,
    FIRST_FRAME_ASPECT_RATIOS,
    MAX_DURATION,
    MEGAPIXELS_MAX,
    MEGAPIXELS_MIN,
    MEGAPIXELS_PRESETS_16X9,
    MIN_DURATION,
    TEXT_ASPECT_RATIOS,
    RunningHubError,
)
from .runninghub import query_task as rh_query_task
from .runninghub import submit_first_frame_task as rh_submit_first_frame_task
from .runninghub import submit_reference_task as rh_submit_reference_task
from .runninghub import submit_text_task as rh_submit_text_task
from .runninghub import upload_media as rh_upload_media
from .server_monitoring import monitoring_summary
from .storage import get_storage, import_remote, safe_key
from .storyboard_options import OPTION_KINDS, load_general_storyboard_options
from .token_usage import add_llm_call_log, add_token_usage

router = APIRouter(prefix="/api/admin", tags=["admin"])
Db = Depends(database_session)


def require_admin(user):
    require_super_admin(user)


def iso(value):
    return value.isoformat() if value else None


async def audit(db, request, user, action, target_type, target_id=None, before=None, after=None):
    db.add(
        AdminOperationLogModel(
            id=f"audit-{uuid.uuid4().hex}",
            admin_user_id=user.id,
            action=action,
            target_type=target_type,
            target_id=target_id,
            before_data=before or {},
            after_data=after or {},
            client_ip=request.client.host if request.client else None,
        )
    )


class ChatComparisonIn(BaseModel):
    system_prompt: str = Field(default="", max_length=12000)
    prompt: str = Field(min_length=1, max_length=30000)
    models: list[str] = Field(min_length=1, max_length=6)
    temperature: float = Field(default=0.2, ge=0, le=2)
    max_tokens: int = Field(default=2048, ge=1, le=8192)


class GeneralOutlineComparisonIn(BaseModel):
    models: list[str] = Field(min_length=1, max_length=6)
    genre: str = Field(min_length=1, max_length=120)
    secondary_category: str = Field(default="", max_length=120)
    tertiary_category: str = Field(default="", max_length=120)
    season: str = Field(default="", max_length=80)
    gender: str = Field(default="", max_length=32)
    age_group: str = Field(default="", max_length=80)
    visual_style: str = Field(default="", max_length=1000)
    empty_shot_count: int = Field(ge=0, le=30)
    character_shot_count: int = Field(ge=0, le=30)
    total_duration: float = Field(gt=0, le=600)
    extra_requirement: str = Field(default="", max_length=4000)
    overall_prompt: str = Field(default="", max_length=8000)
    character_name: str = Field(default="测试人物", min_length=1, max_length=120)
    character_age: str = Field(default="", max_length=120)
    character_appearance: str = Field(default="", max_length=1000)
    character_clothing: str = Field(default="", max_length=1000)


@router.get("/chat-comparison/models")
async def chat_comparison_models(user: CurrentUser):
    require_admin(user)
    return [{"code": item.code, "name": item.name, "protocol": item.protocol} for item in CHAT_TEST_MODELS]


@router.post("/chat-comparison/run")
async def run_chat_comparison(payload: ChatComparisonIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    models = list(dict.fromkeys(payload.models))
    allowed = {item.code for item in CHAT_TEST_MODELS}
    unknown = [model for model in models if model not in allowed]
    if unknown:
        raise HTTPException(422, f"不支持的 Chat 模型：{', '.join(unknown)}")
    messages = ([{"role": "system", "content": payload.system_prompt}] if payload.system_prompt else []) + [{"role": "user", "content": payload.prompt}]
    results = await compare_chat_models(
        models=models,
        system_prompt=payload.system_prompt,
        prompt=payload.prompt,
        temperature=payload.temperature,
        max_tokens=payload.max_tokens,
    )
    for result in results:
        usage = result["usage"].get("raw") or {}
        add_token_usage(
            db,
            operation="admin_chat_comparison",
            provider=result["protocol"],
            model=result["model"],
            usage=usage,
            user_id=user.id,
            request_id=result.get("requestId"),
        )
        add_llm_call_log(
            db,
            operation="admin_chat_comparison",
            provider=result["protocol"],
            model=result["model"],
            usage=usage,
            user_id=user.id,
            request_id=result.get("requestId"),
            status=result["status"],
            error=result["error"],
            duration_ms=result["durationMs"],
            request_messages=messages,
            response_text=result["text"],
        )
    await audit(db, request, user, "chat_comparison.run", "llm_comparison", after={"models": models, "temperature": payload.temperature, "maxTokens": payload.max_tokens})
    await db.commit()
    return {"results": results}


@router.post("/chat-comparison/general-outline")
async def run_general_outline_comparison(payload: GeneralOutlineComparisonIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    if payload.empty_shot_count + payload.character_shot_count < 1:
        raise HTTPException(422, "镜头总数至少为 1")
    models = list(dict.fromkeys(payload.models))
    allowed = {item.code for item in CHAT_TEST_MODELS}
    unknown = [model for model in models if model not in allowed]
    if unknown:
        raise HTTPException(422, f"不支持的 Chat 模型：{', '.join(unknown)}")
    config = payload.model_dump(exclude={"models", "character_name", "character_age", "character_appearance", "character_clothing"})
    selected_humans = []
    if payload.character_shot_count:
        selected_humans = [
            {
                "id": "comparison-character-1",
                "name": payload.character_name,
                "ageDescription": payload.character_age,
                "appearanceStyle": payload.character_appearance,
                "clothingDescription": payload.character_clothing,
            }
        ]
    results = await compare_general_outlines(models=models, config=config, selected_humans=selected_humans)
    for result in results:
        for call in result["calls"]:
            add_token_usage(
                db,
                operation="admin_general_outline_comparison",
                provider=result["protocol"],
                model=result["model"],
                usage=call.get("usage"),
                user_id=user.id,
                request_id=call.get("requestId"),
            )
            add_llm_call_log(
                db,
                operation="admin_general_outline_comparison",
                provider=result["protocol"],
                model=result["model"],
                usage=call.get("usage"),
                user_id=user.id,
                request_id=call.get("requestId"),
                status=call["status"],
                error=call.get("error", ""),
                duration_ms=call["durationMs"],
                request_messages=call["requestMessages"],
                response_text=call["responseText"],
                prompt_key=call.get("promptKey", ""),
                prompt_version=call.get("promptVersion", 0),
            )
    await audit(
        db,
        request,
        user,
        "general_outline_comparison.run",
        "llm_comparison",
        after={"models": models, "emptyShotCount": payload.empty_shot_count, "characterShotCount": payload.character_shot_count},
    )
    await db.commit()
    public_results = [
        {key: value for key, value in result.items() if key != "calls"}
        | {"callMetrics": [{"operation": call["operation"], "status": call["status"], "durationMs": call["durationMs"]} for call in result["calls"]]}
        for result in results
    ]
    return {"results": public_results}


@router.get("/dashboard")
async def dashboard(user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)

    async def count(model, *where):
        return (await db.execute(select(func.count()).select_from(model).where(model.deleted_at.is_(None), *where))).scalar_one()

    usage = (
        await db.execute(
            select(
                func.coalesce(func.sum(TokenUsageModel.input_tokens), 0),
                func.coalesce(func.sum(TokenUsageModel.output_tokens), 0),
                func.coalesce(func.sum(TokenUsageModel.total_tokens), 0),
            ).where(TokenUsageModel.deleted_at.is_(None))
        )
    ).one()
    statuses = dict((await db.execute(select(GenerationJobModel.status, func.count()).where(GenerationJobModel.deleted_at.is_(None)).group_by(GenerationJobModel.status))).all())
    return {
        "users": await count(UserModel),
        "projects": await count(ProjectModel),
        "jobs": await count(GenerationJobModel),
        "systemHumans": await count(DigitalHumanModel, DigitalHumanModel.scope == "system"),
        "errors": await count(ApiErrorLogModel),
        "usage": {"inputTokens": usage[0], "outputTokens": usage[1], "totalTokens": usage[2]},
        "jobStatuses": statuses,
    }


@router.get("/server-monitoring")
async def server_monitoring(user: CurrentUser, hours: int = 24, db: AsyncSession = Db):
    require_admin(user)
    result = await monitoring_summary(db, hours)
    runs = list(
        (await db.execute(select(ServerMaintenanceRunModel).where(ServerMaintenanceRunModel.deleted_at.is_(None)).order_by(ServerMaintenanceRunModel.created_at.desc()).limit(20)))
        .scalars()
        .all()
    )
    result["maintenanceRuns"] = [
        {"id": x.id, "action": x.action, "trigger": x.trigger, "dryRun": x.dry_run, "status": x.status, "summary": x.summary, "details": x.details, "createdAt": iso(x.created_at)}
        for x in runs
    ]
    return result


class ServerMaintenanceDryRunIn(BaseModel):
    action: str = Field(pattern="^(cleanup_temp_files|cleanup_dangling_images|rotate_logs)$")


@router.post("/server-monitoring/maintenance/dry-run", status_code=201)
async def server_maintenance_dry_run(payload: ServerMaintenanceDryRunIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """只生成白名单维护方案并审计；第一阶段不在 Web 进程执行宿主机删除。"""
    require_admin(user)
    labels = {
        "cleanup_temp_files": "扫描专用临时目录中超过 24 小时且已完成任务的文件",
        "cleanup_dangling_images": "扫描 dangling 镜像，保留当前及上一发布版本",
        "rotate_logs": "扫描超过轮转阈值的应用与维护日志",
    }
    item = ServerMaintenanceRunModel(
        id=f"maintenance-{uuid.uuid4().hex}",
        requested_by=user.id,
        source="primary",
        action=payload.action,
        trigger="manual",
        dry_run=True,
        status="completed",
        summary=f"DRY-RUN：{labels[payload.action]}",
        details={"executed": False, "safety": "allowlist-only", "requiresManualExecution": True},
    )
    db.add(item)
    await audit(db, request, user, "server_maintenance.dry_run", "server_maintenance", item.id, None, {"action": payload.action})
    await db.commit()
    return {"id": item.id, "status": item.status, "dryRun": True, "summary": item.summary, "details": item.details}


@router.get("/projects")
async def projects(
    user: CurrentUser,
    limit: int = 100,
    offset: int = 0,
    db: AsyncSession = Db,
):
    """项目列表：offset 分页，limit 上限 300，防止全量拉取拖垮数据库。"""
    require_admin(user)
    limit, offset = max(1, min(limit, 300)), max(0, offset)
    conditions = [ProjectModel.deleted_at.is_(None)]
    total = (await db.execute(select(func.count()).select_from(ProjectModel).where(*conditions))).scalar_one()
    rows = (
        await db.execute(
            select(ProjectModel, UserModel.username)
            .join(UserModel, UserModel.id == ProjectModel.user_id)
            .where(*conditions)
            .order_by(ProjectModel.created_at.desc())
            .limit(limit)
            .offset(offset)
        )
    ).all()
    return {
        "total": total,
        "items": [{"id": p.id, "name": p.name, "username": u, "status": p.status, "createdAt": iso(p.created_at)} for p, u in rows],
    }


@router.get("/jobs")
async def jobs(
    user: CurrentUser,
    db: AsyncSession = Db,
    kind: str | None = None,
    status: str | None = None,
    q: str | None = None,
    page: int = 1,
    page_size: int = 50,
):
    """生成任务全量收集：含供应商 taskId，支持类型/状态/关键词筛选与分页"""
    require_admin(user)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    conditions = [GenerationJobModel.deleted_at.is_(None)]
    if kind in {"image", "video"}:
        conditions.append(GenerationJobModel.kind == kind)
    if status:
        conditions.append(GenerationJobModel.status == status)
    if q and q.strip():
        term = f"%{q.strip()}%"
        conditions.append(or_(GenerationJobModel.id.ilike(term), GenerationJobModel.provider_task_id.ilike(term)))
    total = (await db.execute(select(func.count()).select_from(GenerationJobModel).where(*conditions))).scalar_one()
    stale_cutoff = utcnow() - timedelta(minutes=10)
    stale_col = case(
        (GenerationJobModel.status.in_(("queued", "running")) & (GenerationJobModel.updated_at < stale_cutoff), True),
        else_=False,
    )
    rows = (
        await db.execute(
            select(GenerationJobModel, UserModel.username, stale_col.label("stale"))
            .join(UserModel, UserModel.id == GenerationJobModel.user_id, isouter=True)
            .where(*conditions)
            .order_by(GenerationJobModel.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    ).all()
    items = [
        {
            "id": j.id,
            "username": username,
            "kind": j.kind,
            "status": j.status,
            "progress": j.progress,
            "provider": j.provider,
            "providerTaskId": j.provider_task_id,
            "model": (j.request or {}).get("model"),
            "storyboardLineId": j.storyboard_line_id,
            "error": j.error,
            "stale": bool(stale),
            "durationSeconds": round((j.finished_at - j.started_at).total_seconds()) if j.finished_at and j.started_at else None,
            "createdAt": iso(j.created_at),
            "finishedAt": iso(j.finished_at),
        }
        for j, username, stale in rows
    ]
    return {"total": total, "page": page, "pageSize": page_size, "items": items}


@router.post("/jobs/{job_id}/sync")
async def sync_generation_job(job_id: str, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """按供应商 taskId 主动对账：已成功则补落库资产挽回结果，已失败同步原因，供应商仍在跑而本机无协程则重新挂轮询"""
    require_admin(user)
    model = await db.get(GenerationJobModel, job_id)
    if not model or model.deleted_at is not None:
        raise HTTPException(404, "任务不存在")
    if not model.provider_task_id:
        raise HTTPException(422, "该任务没有供应商任务ID，无法同步")
    if job_manager.is_active(job_id):
        return {"providerStatus": None, "action": "skipped", "detail": "任务正在本机执行中，无需同步"}
    try:
        data = await query_provider_task(model.kind, model.provider_task_id, model.provider)
    except Exception as exc:
        raise HTTPException(502, f"查询供应商失败：{str(exc)[:300]}") from exc
    provider_status = str(data.get("status", "")).upper()
    action = "unchanged"
    if provider_status == "SUCCESS" and model.status != "succeeded":
        job = await job_manager.get(job_id)
        if job:
            result = await store_provider_result(job, data)
            await job_manager.finalize_success(job, result)
            action = "recovered"
    elif (provider_status in {"FAILED", "CANCELLED"} or "FAIL" in provider_status) and model.status != "failed":
        job = await job_manager.get(job_id)
        if job:
            await job_manager.finalize_failure(job, str(data.get("failReason") or f"供应商任务状态：{provider_status}"))
            action = "failed"
    elif "FAIL" not in provider_status and provider_status != "CANCELLED" and model.status in {"queued", "running"}:
        if await job_manager.resume_one(job_id, resume_generation):
            action = "resumed"
    await audit(db, request, user, "job.sync", "generation_job", job_id, after={"providerStatus": provider_status, "action": action})
    await db.commit()
    return {"providerStatus": provider_status, "action": action, "progress": data.get("progress")}


@router.get("/usage")
async def usage(user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    rows = (
        await db.execute(
            select(
                TokenUsageModel.model,
                TokenUsageModel.provider,
                func.sum(TokenUsageModel.input_tokens),
                func.sum(TokenUsageModel.output_tokens),
                func.sum(TokenUsageModel.total_tokens),
                func.count(),
            )
            .where(TokenUsageModel.deleted_at.is_(None))
            .group_by(TokenUsageModel.model, TokenUsageModel.provider)
        )
    ).all()
    return [{"model": r[0], "provider": r[1], "inputTokens": r[2] or 0, "outputTokens": r[3] or 0, "totalTokens": r[4] or 0, "calls": r[5]} for r in rows]


class ProviderIn(BaseModel):
    code: str = Field(min_length=1, max_length=80)
    name: str = Field(min_length=1, max_length=120)
    base_url: str = ""
    status: str = "active"


class ModelIn(BaseModel):
    provider_id: str
    code: str = Field(min_length=1, max_length=160)
    name: str
    modality: str
    provider_model_id: str
    capabilities: dict = {}
    status: str = "active"
    user_visible: bool = True
    is_default: bool = False


@router.get("/providers")
async def providers(user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    rows = (await db.execute(select(AiProviderModel).where(AiProviderModel.deleted_at.is_(None)).order_by(AiProviderModel.name))).scalars()
    return [{"id": x.id, "code": x.code, "name": x.name, "baseUrl": x.base_url, "status": x.status} for x in rows]


@router.post("/providers", status_code=201)
async def create_provider(payload: ProviderIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    item = AiProviderModel(id=f"provider-{uuid.uuid4().hex}", code=payload.code, name=payload.name, base_url=payload.base_url, status=payload.status)
    db.add(item)
    await audit(db, request, user, "provider.create", "ai_provider", item.id, after=payload.model_dump())
    await db.commit()
    return {"id": item.id}


@router.get("/models")
async def models(user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    rows = (await db.execute(select(AiModelModel).where(AiModelModel.deleted_at.is_(None)).order_by(AiModelModel.modality, AiModelModel.sort_order))).scalars()
    return [
        {
            "id": x.id,
            "providerId": x.provider_id,
            "code": x.code,
            "name": x.name,
            "modality": x.modality,
            "providerModelId": x.provider_model_id,
            "capabilities": x.capabilities,
            "status": x.status,
            "userVisible": x.user_visible,
            "isDefault": x.is_default,
        }
        for x in rows
    ]


@router.post("/models", status_code=201)
async def create_model(payload: ModelIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    item = AiModelModel(id=f"model-{uuid.uuid4().hex}", **payload.model_dump())
    db.add(item)
    await audit(db, request, user, "model.create", "ai_model", item.id, after=payload.model_dump())
    await db.commit()
    return {"id": item.id}


@router.patch("/models/{model_id}")
async def update_model(model_id: str, payload: dict, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_admin(user)
    item = await db.get(AiModelModel, model_id)
    if not item or item.deleted_at:
        raise HTTPException(404, "模型不存在")
    allowed = {"name", "status", "user_visible", "is_default", "capabilities", "sort_order"}
    before = {k: getattr(item, k) for k in allowed}
    for k, v in payload.items():
        if k in allowed:
            setattr(item, k, v)
    await audit(db, request, user, "model.update", "ai_model", item.id, before, payload)
    await db.commit()
    return {"ok": True}


@router.get("/audit-logs")
async def audits(
    user: CurrentUser,
    limit: int = 100,
    offset: int = 0,
    db: AsyncSession = Db,
):
    """操作审计列表：offset 分页，limit 上限 300，防止全量拉取拖垮数据库。"""
    require_admin(user)
    limit, offset = max(1, min(limit, 300)), max(0, offset)
    conditions = [AdminOperationLogModel.deleted_at.is_(None)]
    total = (await db.execute(select(func.count()).select_from(AdminOperationLogModel).where(*conditions))).scalar_one()
    rows = (await db.execute(select(AdminOperationLogModel).where(*conditions).order_by(AdminOperationLogModel.created_at.desc()).limit(limit).offset(offset))).scalars()
    return {
        "total": total,
        "items": [
            {
                "id": x.id,
                "adminUserId": x.admin_user_id,
                "action": x.action,
                "targetType": x.target_type,
                "targetId": x.target_id,
                "createdAt": iso(x.created_at),
            }
            for x in rows
        ],
    }


def _llm_call_summary(x: LlmCallLogModel) -> dict:
    return {
        "id": x.id,
        "operation": x.operation,
        "model": x.model,
        "status": x.status,
        "error": x.error,
        "durationMs": x.duration_ms,
        "inputTokens": x.input_tokens,
        "outputTokens": x.output_tokens,
        "cachedInputTokens": x.cached_input_tokens,
        "totalTokens": x.total_tokens,
        "requestId": x.request_id,
        "userId": x.user_id,
        "projectId": x.project_id,
        "projectTaskId": x.project_task_id,
        "storyboardLineId": x.storyboard_line_id,
        "generationJobId": x.generation_job_id,
        "promptKey": x.prompt_key,
        "promptVersion": x.prompt_version,
        "createdAt": iso(x.created_at),
    }


@router.get("/llm-calls")
async def llm_calls(
    user: CurrentUser,
    projectTaskId: str | None = None,
    storyboardLineId: str | None = None,
    operation: str | None = None,
    status: str | None = None,
    requestId: str | None = None,
    limit: int = 50,
    offset: int = 0,
    db: AsyncSession = Db,
):
    """分镜 LLM 调用留痕列表（不含请求/返回原文，详情走 /llm-calls/{id}）。"""
    require_admin(user)
    limit, offset = max(1, min(limit, 200)), max(0, offset)
    conditions = [LlmCallLogModel.deleted_at.is_(None)]
    if projectTaskId:
        conditions.append(LlmCallLogModel.project_task_id == projectTaskId)
    if storyboardLineId:
        conditions.append(LlmCallLogModel.storyboard_line_id == storyboardLineId)
    if operation:
        conditions.append(LlmCallLogModel.operation == operation)
    if status:
        conditions.append(LlmCallLogModel.status == status)
    if requestId:
        conditions.append(LlmCallLogModel.request_id == requestId)
    total = (await db.execute(select(func.count()).select_from(LlmCallLogModel).where(*conditions))).scalar_one()
    rows = (await db.execute(select(LlmCallLogModel).where(*conditions).order_by(LlmCallLogModel.created_at.desc()).limit(limit).offset(offset))).scalars()
    return {"total": total, "items": [_llm_call_summary(x) for x in rows]}


@router.get("/llm-calls/{log_id}")
async def llm_call_detail(log_id: str, user: CurrentUser, db: AsyncSession = Db):
    """单条 LLM 调用的全量详情：含请求消息快照与返回原文。"""
    require_admin(user)
    item = await db.get(LlmCallLogModel, log_id)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "调用记录不存在")
    return {**_llm_call_summary(item), "provider": item.provider, "requestMessages": item.request_messages or [], "responseText": item.response_text or ""}


def _request_log_summary(x: ApiRequestLogModel) -> dict:
    return {
        "id": x.id,
        "runId": x.run_id,
        "method": x.method,
        "path": x.path,
        "queryString": x.query_string,
        "statusCode": x.status_code,
        "durationMs": x.duration_ms,
        "userId": x.user_id,
        "clientIp": x.client_ip,
        "createdAt": iso(x.created_at),
    }


@router.get("/request-logs/runs")
async def request_log_runs(user: CurrentUser, db: AsyncSession = Db):
    """测试批次列表：每个批次对应一次全量测试，附耗时统计。"""
    require_admin(user)
    rows = (
        await db.execute(
            select(
                ApiRequestLogModel.run_id,
                func.count().label("requests"),
                func.avg(ApiRequestLogModel.duration_ms).label("avg_ms"),
                func.max(ApiRequestLogModel.duration_ms).label("max_ms"),
                func.min(ApiRequestLogModel.created_at).label("started_at"),
                func.max(ApiRequestLogModel.created_at).label("finished_at"),
                func.sum(case((ApiRequestLogModel.status_code >= 500, 1), else_=0)).label("errors"),
            )
            .where(ApiRequestLogModel.deleted_at.is_(None), ApiRequestLogModel.run_id != "")
            .group_by(ApiRequestLogModel.run_id)
            .order_by(func.max(ApiRequestLogModel.created_at).desc())
            .limit(50)
        )
    ).all()
    return [
        {
            "runId": row.run_id,
            "requests": row.requests,
            "avgMs": round(row.avg_ms or 0),
            "maxMs": row.max_ms or 0,
            "errors": row.errors or 0,
            "startedAt": iso(row.started_at),
            "finishedAt": iso(row.finished_at),
        }
        for row in rows
    ]


@router.get("/request-logs")
async def request_logs(
    user: CurrentUser,
    runId: str | None = None,
    path: str | None = None,
    method: str | None = None,
    status: int | None = None,
    minMs: int | None = None,
    orderBy: str = "created",
    limit: int = 100,
    offset: int = 0,
    db: AsyncSession = Db,
):
    """API 请求耗时列表（不含输入输出原文，详情走 /request-logs/{id}）。

    minMs=仅返回耗时不低于该值的请求；orderBy=duration 时按耗时倒序（慢请求 TOP）。
    """
    require_admin(user)
    limit, offset = max(1, min(limit, 200)), max(0, offset)
    conditions = [ApiRequestLogModel.deleted_at.is_(None)]
    if runId:
        conditions.append(ApiRequestLogModel.run_id == runId)
    if path:
        conditions.append(ApiRequestLogModel.path.contains(path))
    if method:
        conditions.append(ApiRequestLogModel.method == method.upper())
    if status:
        conditions.append(ApiRequestLogModel.status_code == status)
    if minMs is not None:
        conditions.append(ApiRequestLogModel.duration_ms >= max(0, minMs))
    order_by_clause = (ApiRequestLogModel.duration_ms.desc(), ApiRequestLogModel.created_at.desc()) if orderBy == "duration" else (ApiRequestLogModel.created_at.desc(),)
    total = (await db.execute(select(func.count()).select_from(ApiRequestLogModel).where(*conditions))).scalar_one()
    rows = (await db.execute(select(ApiRequestLogModel).where(*conditions).order_by(*order_by_clause).limit(limit).offset(offset))).scalars()
    return {"total": total, "items": [_request_log_summary(x) for x in rows]}


@router.get("/request-logs/summary")
async def request_log_summary(
    user: CurrentUser,
    db: AsyncSession = Db,
    hours: int = 24,
    minCount: int = 3,
    limit: int = 50,
):
    """按 path+method 聚合正式流量（run_id 为空）的耗时分布：count/avg/p95/max。

    直接回答「哪个接口慢」：按 max 倒序取 TOP；p95 精确计算识别偶发慢请求
    与稳定慢接口。跨方言实现（p95 在 Python 计算，不依赖 PostgreSQL
    percentile_cont，测试库 SQLite 同样可用）。
    """
    require_admin(user)
    hours = max(1, min(hours, 168))
    limit = max(1, min(limit, 100))
    since = utcnow() - timedelta(hours=hours)
    # 取时间窗口内最近 2 万条正式流量的耗时明细（防止极端流量下内存膨胀），
    # 在 Python 中按 path+method 分桶计算 count/avg/p95/max
    rows = (
        await db.execute(
            select(
                ApiRequestLogModel.path,
                ApiRequestLogModel.method,
                ApiRequestLogModel.duration_ms,
            )
            .where(
                ApiRequestLogModel.deleted_at.is_(None),
                ApiRequestLogModel.created_at >= since,
                ApiRequestLogModel.run_id == "",
            )
            .order_by(ApiRequestLogModel.created_at.desc())
            .limit(20_000)
        )
    ).all()
    buckets: dict[tuple[str, str], list[int]] = {}
    for path, method, duration_ms in rows:
        buckets.setdefault((path, method), []).append(duration_ms)
    summary = []
    for (path, method), durations in buckets.items():
        if len(durations) < minCount:
            continue
        durations.sort()
        summary.append(
            {
                "path": path,
                "method": method,
                "count": len(durations),
                "avgMs": round(sum(durations) / len(durations)),
                "p95Ms": durations[int(len(durations) * 0.95) - 1],
                "maxMs": durations[-1],
            }
        )
    summary.sort(key=lambda row: row["maxMs"], reverse=True)
    return summary[:limit]


@router.get("/request-logs/{log_id}")
async def request_log_detail(log_id: str, user: CurrentUser, db: AsyncSession = Db):
    """单条请求的全量详情：含脱敏后的输入参数与输出原文。"""
    require_admin(user)
    item = await db.get(ApiRequestLogModel, log_id)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "请求日志不存在")
    return {**_request_log_summary(item), "requestPayload": item.request_payload or {}, "responseBody": item.response_body or {}}


public_router = APIRouter(prefix="/api")


# ── 提示词注册中心（版本化 + 发布/回滚 + 试渲染） ─────────────────────────────


class PromptDraftIn(BaseModel):
    content: str = Field(min_length=1)
    change_note: str = ""


class PromptPublishIn(BaseModel):
    version_id: str


class PromptPreviewIn(BaseModel):
    content: str
    variables: dict[str, Any] = {}


def _prompt_version_summary(v: PromptVersionModel) -> dict:
    return {
        "id": v.id,
        "version": v.version,
        "status": v.status,
        "changeNote": v.change_note,
        "createdBy": v.created_by,
        "publishedAt": iso(v.published_at),
        "createdAt": iso(v.created_at),
    }


async def _get_prompt_template(db: AsyncSession, key: str) -> PromptTemplateModel:
    template = (await db.execute(select(PromptTemplateModel).where(PromptTemplateModel.key == key, PromptTemplateModel.deleted_at.is_(None)))).scalar_one_or_none()
    if not template:
        raise HTTPException(404, "提示词模板不存在")
    return template


def _validate_prompt_content(template: PromptTemplateModel, content: str) -> None:
    """发布前校验：安全片段必含、模板变量必须已声明、json 模板必须是字符串数组。"""
    missing = [fragment for fragment in (template.required_fragments or []) if fragment not in content]
    if missing:
        raise HTTPException(422, f"缺少必含安全片段：{missing}")
    undeclared = [name for name in template_variables(content) if name not in (template.variables or {})]
    if undeclared:
        raise HTTPException(422, f"模板变量未声明（代码不会为其传值）：{undeclared}")
    if (template.format or "text") == "json":
        try:
            parsed = json.loads(content)
        except ValueError as exc:
            raise HTTPException(422, f"JSON 模板解析失败：{exc}") from exc
        if not isinstance(parsed, list) or not all(isinstance(item, str) for item in parsed):
            raise HTTPException(422, "JSON 模板必须是字符串数组")


@router.get("/prompts")
async def prompt_templates(user: CurrentUser, db: AsyncSession = Db):
    """提示词模板列表：含当前已发布版本摘要。"""
    require_admin(user)
    rows = (await db.execute(select(PromptTemplateModel).where(PromptTemplateModel.deleted_at.is_(None)).order_by(PromptTemplateModel.engine, PromptTemplateModel.key))).scalars()
    result = []
    for item in rows:
        current = await db.get(PromptVersionModel, item.current_version_id) if item.current_version_id else None
        result.append(
            {
                "id": item.id,
                "key": item.key,
                "name": item.name,
                "description": item.description,
                "engine": item.engine,
                "format": item.format,
                "variables": item.variables or {},
                "requiredFragments": item.required_fragments or [],
                "currentVersion": _prompt_version_summary(current) if current and current.deleted_at is None else None,
                "updatedAt": iso(item.updated_at),
            }
        )
    return result


@router.get("/prompts/{key}")
async def prompt_template_detail(key: str, user: CurrentUser, db: AsyncSession = Db):
    """模板详情：全部版本（新到旧）+ 内置默认内容（用于 diff 与一键恢复）。"""
    require_admin(user)
    template = await _get_prompt_template(db, key)
    versions = (
        await db.execute(
            select(PromptVersionModel).where(PromptVersionModel.template_id == template.id, PromptVersionModel.deleted_at.is_(None)).order_by(PromptVersionModel.version.desc())
        )
    ).scalars()
    default = DEFAULT_PROMPTS.get(key)
    return {
        "id": template.id,
        "key": template.key,
        "name": template.name,
        "description": template.description,
        "engine": template.engine,
        "format": template.format or "text",
        "variables": template.variables or {},
        "requiredFragments": template.required_fragments or [],
        "currentVersionId": template.current_version_id,
        "versions": [{**_prompt_version_summary(v), "content": v.content} for v in versions],
        "defaultContent": default["content"] if default else None,
    }


@router.post("/prompts/{key}/versions", status_code=201)
async def prompt_create_draft(key: str, payload: PromptDraftIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """新建草稿版本（版本号递增）；内容校验在发布时才强制执行，草稿允许暂存不完整内容。"""
    require_admin(user)
    template = await _get_prompt_template(db, key)
    latest = (await db.execute(select(func.max(PromptVersionModel.version)).where(PromptVersionModel.template_id == template.id))).scalar_one()
    draft = PromptVersionModel(
        id=f"pv-{uuid.uuid4().hex}",
        template_id=template.id,
        version=int(latest or 0) + 1,
        content=payload.content,
        change_note=payload.change_note,
        status="draft",
        created_by=getattr(user, "username", "") or user.id,
    )
    db.add(draft)
    await audit(db, request, user, "prompt.draft", "prompt_template", template.id, after={"key": key, "version": draft.version, "changeNote": payload.change_note})
    await db.commit()
    return {"id": draft.id, "version": draft.version}


@router.post("/prompts/{key}/publish")
async def prompt_publish(key: str, payload: PromptPublishIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """发布指定版本；回滚 = 选择旧版本发布。发布前强制校验，发布后注册中心缓存立即失效。"""
    require_admin(user)
    template = await _get_prompt_template(db, key)
    version = await db.get(PromptVersionModel, payload.version_id)
    if not version or version.template_id != template.id or version.deleted_at is not None:
        raise HTTPException(404, "提示词版本不存在")
    _validate_prompt_content(template, version.content)
    before = {"currentVersionId": template.current_version_id}
    current = await db.get(PromptVersionModel, template.current_version_id) if template.current_version_id else None
    if current and current.status == "published":
        current.status = "archived"
    version.status = "published"
    version.published_at = utcnow()
    template.current_version_id = version.id
    await audit(db, request, user, "prompt.publish", "prompt_template", template.id, before, {"currentVersionId": version.id, "version": version.version})
    await db.commit()
    invalidate(key)
    return {"ok": True, "version": version.version}


@router.post("/prompts/{key}/preview")
async def prompt_preview(key: str, payload: PromptPreviewIn, user: CurrentUser, db: AsyncSession = Db):
    """试渲染：只替换已提供的变量，返回渲染结果与校验报告，不调用任何模型。"""
    require_admin(user)
    template = await _get_prompt_template(db, key)
    used = template_variables(payload.content)
    rendered = render_lenient(payload.content, payload.variables)
    json_error = ""
    if (template.format or "text") == "json":
        try:
            parsed = json.loads(rendered)
            if not isinstance(parsed, list) or not all(isinstance(item, str) for item in parsed):
                json_error = "JSON 模板必须是字符串数组"
        except ValueError as exc:
            json_error = str(exc)
    return {
        "rendered": rendered,
        "usedVariables": used,
        "missingVariables": [name for name in used if name not in payload.variables],
        "undeclaredVariables": [name for name in used if name not in (template.variables or {})],
        "missingFragments": [fragment for fragment in (template.required_fragments or []) if fragment not in payload.content],
        "jsonError": json_error,
    }


@router.delete("/prompts/{key}/versions/{version_id}")
async def prompt_delete_draft(key: str, version_id: str, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """软删除草稿版本；published/archived 版本保留历史不可删除。"""
    require_admin(user)
    template = await _get_prompt_template(db, key)
    version = await db.get(PromptVersionModel, version_id)
    if not version or version.template_id != template.id or version.deleted_at is not None:
        raise HTTPException(404, "提示词版本不存在")
    if version.status != "draft":
        raise HTTPException(409, "仅草稿版本可删除")
    version.deleted_at = utcnow()
    await audit(db, request, user, "prompt.discard", "prompt_template", template.id, {"version": version.version})
    await db.commit()
    return {"ok": True}


@public_router.get("/model-options")
async def model_options(user: CurrentUser, modality: str | None = None, db: AsyncSession = Db):
    query = select(AiModelModel).where(AiModelModel.deleted_at.is_(None), AiModelModel.status == "active", AiModelModel.user_visible.is_(True))
    if modality:
        query = query.where(AiModelModel.modality == modality)
    rows = (await db.execute(query.order_by(AiModelModel.modality, AiModelModel.sort_order))).scalars()
    return [{"id": x.code, "name": x.name, "modality": x.modality, "capabilities": x.capabilities, "isDefault": x.is_default} for x in rows]


@public_router.get("/aigc/models")
async def aigc_provider_models(user: CurrentUser):
    """工具：实时查询上游 AIGC 平台当前账号可见的模型列表（/v1/models）。

    用途：核对环境变量（VIDEO_API_KEY/AIGC_TOKEN）指向的账号在平台侧实际开放了哪些
    模型，排查“模型名不存在/不可用”类问题；key 与生成链路同一来源，不做缓存。
    """
    try:
        models = await list_video_models()
    except ProviderError as exc:
        raise HTTPException(502, f"上游模型列表查询失败：{exc}") from exc
    return {
        "provider": "yinghe",
        "baseUrl": settings.video_api_base_url,
        "apiKeySuffix": settings.video_api_key[-6:],
        "count": len(models),
        "models": models,
    }


# ---------- ASS 歌曲情感库 ----------


class SongEmotionProfileIn(BaseModel):
    song_code: str = Field(min_length=5, max_length=80, pattern=r"^\d+$")
    song_name: str = Field(default="", max_length=255)
    artists: str = Field(default="", max_length=2000)
    lyrics: str = Field(default="", max_length=100000)
    primary_category: str | None = Field(default=None, max_length=255)
    secondary_category: str | None = Field(default=None, max_length=255)
    tertiary_category: str | None = Field(default=None, max_length=255)
    material_category: str = Field(default="", max_length=2000)
    seasons: str = Field(default="", max_length=120)
    atmosphere: str = Field(default="", max_length=8000)
    character_setting: str = Field(default="", max_length=8000)
    status: int = 2


class SongEmotionProfilePatch(BaseModel):
    song_name: str | None = Field(default=None, max_length=255)
    artists: str | None = Field(default=None, max_length=2000)
    lyrics: str | None = Field(default=None, max_length=100000)
    primary_category: str | None = Field(default=None, max_length=255)
    secondary_category: str | None = Field(default=None, max_length=255)
    tertiary_category: str | None = Field(default=None, max_length=255)
    material_category: str | None = Field(default=None, max_length=2000)
    seasons: str | None = Field(default=None, max_length=120)
    atmosphere: str | None = Field(default=None, max_length=8000)
    character_setting: str | None = Field(default=None, max_length=8000)
    status: int | None = None


def _clean_optional(value: str | None) -> str | None:
    return (value.strip() or None) if value is not None else None


def _song_payload(item: SongEmotionProfileModel) -> dict:
    return {
        "歌名": item.song_name,
        "歌星": item.artists,
        "歌词": item.lyrics,
        "一级分类": item.primary_category,
        "二级分类": item.secondary_category,
        "三级分类": item.tertiary_category,
        "素材分类": item.material_category,
        "季节": item.seasons,
        "氛围基调": item.atmosphere,
        "人物设定": item.character_setting,
        "状态": item.status,
    }


def _song_summary(item: SongEmotionProfileModel) -> dict:
    return {
        "songCode": item.song_code,
        "songName": item.song_name,
        "artists": item.artists,
        "lyrics": item.lyrics,
        "primaryCategory": item.primary_category,
        "secondaryCategory": item.secondary_category,
        "tertiaryCategory": item.tertiary_category,
        "materialCategory": item.material_category,
        "seasons": item.seasons,
        "atmosphere": item.atmosphere,
        "characterSetting": item.character_setting,
        "status": item.status,
        "createdAt": iso(item.created_at),
        "updatedAt": iso(item.updated_at),
    }


def _apply_song_values(item: SongEmotionProfileModel, values: dict) -> None:
    for key, value in values.items():
        if key in {"primary_category", "secondary_category", "tertiary_category"}:
            setattr(item, key, _clean_optional(value))
        elif isinstance(value, str):
            setattr(item, key, value.strip())
    item.source_payload = _song_payload(item)


async def _validate_song_taxonomy(db: AsyncSession, item: SongEmotionProfileModel, options: dict | None = None) -> None:
    options = options or await load_general_storyboard_options(db)
    primary = next((x for x in options["genres"] if x["value"] == item.primary_category), None)
    if not primary:
        raise HTTPException(422, "请选择有效的一级分类")
    secondary = None
    if item.secondary_category:
        secondary = next((x for x in primary.get("children", []) if x["value"] == item.secondary_category), None)
        if not secondary:
            raise HTTPException(422, "二级分类不属于所选一级分类")
    if item.tertiary_category:
        tertiary = next((x for x in (secondary or {}).get("children", []) if x["value"] == item.tertiary_category), None)
        if not tertiary:
            raise HTTPException(422, "三级分类不属于所选二级分类")
    seasons = [value for value in item.seasons.split("/") if value]
    if not seasons or any(value not in options["seasons"] for value in seasons) or ("通用" in seasons and len(seasons) > 1):
        raise HTTPException(422, "请选择有效的适用季节")
    item.material_category = "-".join(value for value in (item.primary_category, item.secondary_category, item.tertiary_category) if value)
    item.source_payload = _song_payload(item)


@router.get("/song-emotion-profiles")
async def song_emotion_profiles_list(
    user: CurrentUser,
    db: AsyncSession = Db,
    q: str = "",
    primary_category: str = "",
    secondary_category: str = "",
    tertiary_category: str = "",
    limit: int = 50,
    offset: int = 0,
):
    require_permission(user, SONG_EMOTIONS_READ)
    limit = min(max(limit, 1), 200)
    offset = max(offset, 0)
    filters = [SongEmotionProfileModel.deleted_at.is_(None)]
    if q.strip():
        pattern = f"%{q.strip()}%"
        filters.append(
            or_(
                SongEmotionProfileModel.song_code.ilike(pattern),
                SongEmotionProfileModel.song_name.ilike(pattern),
                SongEmotionProfileModel.artists.ilike(pattern),
            )
        )
    for column, value in (
        (SongEmotionProfileModel.primary_category, primary_category),
        (SongEmotionProfileModel.secondary_category, secondary_category),
        (SongEmotionProfileModel.tertiary_category, tertiary_category),
    ):
        if value.strip():
            filters.append(column == value.strip())
    total = int((await db.execute(select(func.count()).select_from(SongEmotionProfileModel).where(*filters))).scalar_one())
    rows = list((await db.execute(select(SongEmotionProfileModel).where(*filters).order_by(SongEmotionProfileModel.song_code.desc()).limit(limit).offset(offset))).scalars().all())
    return {"total": total, "items": [_song_summary(item) for item in rows]}


SONG_EMOTION_XLSX_HEADERS = [
    "编号",
    "歌名",
    "歌星",
    "歌词",
    "一级分类",
    "二级分类",
    "三级分类",
    "素材分类",
    "季节",
    "氛围基调",
    "人物设定",
    "状态",
]


def _xlsx_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@router.post("/song-emotion-profiles/import-xlsx")
async def song_emotion_profiles_import(
    request: Request,
    user: CurrentUser,
    db: AsyncSession = Db,
    file: UploadFile = File(...),
):
    require_permission(user, SONG_EMOTIONS_MANAGE)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(422, "仅支持 .xlsx 文件")
    content = await file.read(10 * 1024 * 1024 + 1)
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "XLSX 文件不能超过 10MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [_xlsx_text(value) for value in next(rows)]
    except (BadZipFile, InvalidFileException, OSError, ValueError, StopIteration) as exc:
        raise HTTPException(422, "无法读取 XLSX 文件或文件内容为空") from exc
    if headers != SONG_EMOTION_XLSX_HEADERS:
        raise HTTPException(422, f"表头必须依次为：{'、'.join(SONG_EMOTION_XLSX_HEADERS)}")

    records: list[tuple[int, dict[str, str]]] = []
    for row_number, values in enumerate(rows, start=2):
        cells = list(values[: len(headers)])
        if not any(value is not None and _xlsx_text(value) for value in cells):
            continue
        cells.extend([None] * (len(headers) - len(cells)))
        records.append((row_number, dict(zip(headers, map(_xlsx_text, cells), strict=True))))
    workbook.close()
    if not records:
        raise HTTPException(422, "XLSX 中没有可导入的数据")

    errors: list[str] = []
    codes: list[str] = []
    seen: set[str] = set()
    for row_number, record in records:
        code = record["编号"]
        if not code.isdigit() or len(code) < 5:
            errors.append(f"第 {row_number} 行：编号必须是至少5位数字")
        elif code in seen:
            errors.append(f"第 {row_number} 行：编号 {code} 在文件内重复")
        else:
            seen.add(code)
            codes.append(code)
        if not record["歌名"]:
            errors.append(f"第 {row_number} 行：歌名不能为空")

    existing = set((await db.execute(select(SongEmotionProfileModel.song_code).where(SongEmotionProfileModel.song_code.in_(codes)))).scalars().all())
    if existing:
        errors.append(f"以下编号已存在于数据库：{', '.join(sorted(existing))}")
    if errors:
        raise HTTPException(409, "；".join(errors[:20]))

    options = await load_general_storyboard_options(db)
    imported: list[SongEmotionProfileModel] = []
    for row_number, record in records:
        try:
            status = int(record["状态"] or "2")
        except ValueError as exc:
            raise HTTPException(422, f"第 {row_number} 行：状态必须是整数") from exc
        item = SongEmotionProfileModel(
            song_code=record["编号"],
            song_name=record["歌名"],
            artists=record["歌星"],
            lyrics=record["歌词"],
            primary_category=_clean_optional(record["一级分类"]),
            secondary_category=_clean_optional(record["二级分类"]),
            tertiary_category=_clean_optional(record["三级分类"]),
            material_category=record["素材分类"],
            seasons=record["季节"],
            atmosphere=record["氛围基调"],
            character_setting=record["人物设定"],
            status=status,
        )
        try:
            await _validate_song_taxonomy(db, item, options)
        except HTTPException as exc:
            raise HTTPException(422, f"第 {row_number} 行：{exc.detail}") from exc
        imported.append(item)

    db.add_all(imported)
    await db.flush()
    await audit(
        db,
        request,
        user,
        "song_emotion_profile.import",
        "song_emotion_profile",
        file.filename,
        None,
        {"count": len(imported), "songCodes": [item.song_code for item in imported]},
    )
    await db.commit()
    return {"ok": True, "imported": len(imported)}


@router.get("/song-emotion-profiles/{song_code}")
async def song_emotion_profile_detail(song_code: str, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, SONG_EMOTIONS_READ)
    item = await db.get(SongEmotionProfileModel, song_code)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "歌曲情感档案不存在")
    return _song_summary(item)


@router.post("/song-emotion-profiles", status_code=201)
async def song_emotion_profile_create(payload: SongEmotionProfileIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, SONG_EMOTIONS_MANAGE)
    code = payload.song_code.strip()
    existing = await db.get(SongEmotionProfileModel, code)
    if existing:
        raise HTTPException(409, "歌曲编号已存在（包括已删除记录）")
    item = SongEmotionProfileModel(song_code=code)
    _apply_song_values(item, payload.model_dump(exclude={"song_code"}))
    await _validate_song_taxonomy(db, item)
    db.add(item)
    await db.flush()
    await audit(db, request, user, "song_emotion_profile.create", "song_emotion_profile", code, None, _song_summary(item))
    await db.commit()
    return _song_summary(item)


@router.patch("/song-emotion-profiles/{song_code}")
async def song_emotion_profile_update(song_code: str, payload: SongEmotionProfilePatch, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, SONG_EMOTIONS_MANAGE)
    item = await db.get(SongEmotionProfileModel, song_code)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "歌曲情感档案不存在")
    before = _song_summary(item)
    _apply_song_values(item, payload.model_dump(exclude_unset=True))
    await _validate_song_taxonomy(db, item)
    await db.flush()
    await audit(db, request, user, "song_emotion_profile.update", "song_emotion_profile", song_code, before, _song_summary(item))
    await db.commit()
    return _song_summary(item)


@router.delete("/song-emotion-profiles/{song_code}")
async def song_emotion_profile_delete(song_code: str, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, SONG_EMOTIONS_MANAGE)
    item = await db.get(SongEmotionProfileModel, song_code)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "歌曲情感档案不存在")
    before = _song_summary(item)
    item.deleted_at = utcnow()
    await audit(db, request, user, "song_emotion_profile.delete", "song_emotion_profile", song_code, before, None)
    await db.commit()
    return {"ok": True}


# ---------- 通用分镜选项（曲风分类树 / 季节 / 年龄段 / 画面风格） ----------


class StoryboardOptionIn(BaseModel):
    kind: str
    parent_id: str | None = None
    name: str
    sort_order: int | None = None


class StoryboardOptionPatch(BaseModel):
    name: str | None = None
    sort_order: int | None = None


def _option_summary(x: StoryboardOptionItemModel) -> dict:
    return {"id": x.id, "kind": x.kind, "parentId": x.parent_id, "name": x.name, "sortOrder": x.sort_order}


def _validate_option_name(name: str) -> str:
    value = name.strip()
    if not value:
        raise HTTPException(422, "名称不能为空")
    if len(value) > 60:
        raise HTTPException(422, "名称过长（最多 60 字）")
    return value


async def _option_depth(db: AsyncSession, item: StoryboardOptionItemModel) -> int:
    depth = 1
    current = item
    while current.parent_id:
        parent = await db.get(StoryboardOptionItemModel, current.parent_id)
        if not parent or parent.deleted_at is not None:
            break
        depth += 1
        current = parent
    return depth


def _sibling_where(kind: str, parent_id: str | None):
    base = StoryboardOptionItemModel.parent_id.is_(None) if parent_id is None else StoryboardOptionItemModel.parent_id == parent_id
    return (
        StoryboardOptionItemModel.kind == kind,
        base,
        StoryboardOptionItemModel.deleted_at.is_(None),
    )


@router.get("/storyboard-options")
async def storyboard_options_list(kind: str, user: CurrentUser, db: AsyncSession = Db):
    """平铺列表（未删除），前端按 parentId 组树；管理端需要看到完整结构而非仅树。"""
    require_permission(user, STORYBOARD_OPTIONS_READ)
    if kind not in OPTION_KINDS:
        raise HTTPException(422, "未知的选项类型")
    rows = (
        (
            await db.execute(
                select(StoryboardOptionItemModel)
                .where(StoryboardOptionItemModel.kind == kind, StoryboardOptionItemModel.deleted_at.is_(None))
                .order_by(StoryboardOptionItemModel.sort_order, StoryboardOptionItemModel.created_at)
            )
        )
        .scalars()
        .all()
    )
    return [_option_summary(x) for x in rows]


@router.post("/storyboard-options", status_code=201)
async def storyboard_option_create(payload: StoryboardOptionIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, STORYBOARD_OPTIONS_MANAGE)
    if payload.kind not in OPTION_KINDS:
        raise HTTPException(422, "未知的选项类型")
    name = _validate_option_name(payload.name)
    parent = None
    if payload.parent_id:
        if payload.kind != "genre":
            raise HTTPException(422, "仅曲风分类支持子级")
        parent = await db.get(StoryboardOptionItemModel, payload.parent_id)
        if not parent or parent.kind != "genre" or parent.deleted_at is not None:
            raise HTTPException(422, "上级分类不存在")
        if await _option_depth(db, parent) >= 3:
            raise HTTPException(422, "曲风分类最多三级")
    parent_id = parent.id if parent else None
    duplicate = (await db.execute(select(StoryboardOptionItemModel).where(*_sibling_where(payload.kind, parent_id), StoryboardOptionItemModel.name == name))).scalar_one_or_none()
    if duplicate:
        raise HTTPException(409, "同级已存在同名选项")
    if payload.sort_order is not None:
        sort_order = payload.sort_order
    else:
        sort_order = (
            int((await db.execute(select(func.coalesce(func.max(StoryboardOptionItemModel.sort_order), -1)).where(*_sibling_where(payload.kind, parent_id)))).scalar_one()) + 1
        )
    item = StoryboardOptionItemModel(id=f"soi-{uuid.uuid4().hex[:16]}", kind=payload.kind, parent_id=parent_id, name=name, sort_order=sort_order)
    db.add(item)
    await audit(db, request, user, "storyboard_option.create", "storyboard_option_item", item.id, None, _option_summary(item))
    await db.commit()
    return _option_summary(item)


@router.patch("/storyboard-options/{item_id}")
async def storyboard_option_update(item_id: str, payload: StoryboardOptionPatch, request: Request, user: CurrentUser, db: AsyncSession = Db):
    require_permission(user, STORYBOARD_OPTIONS_MANAGE)
    item = await db.get(StoryboardOptionItemModel, item_id)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "选项不存在")
    before = _option_summary(item)
    if payload.name is not None:
        name = _validate_option_name(payload.name)
        duplicate = (
            await db.execute(
                select(StoryboardOptionItemModel).where(
                    *_sibling_where(item.kind, item.parent_id),
                    StoryboardOptionItemModel.name == name,
                    StoryboardOptionItemModel.id != item.id,
                )
            )
        ).scalar_one_or_none()
        if duplicate:
            raise HTTPException(409, "同级已存在同名选项")
        item.name = name
    if payload.sort_order is not None:
        item.sort_order = payload.sort_order
    await audit(db, request, user, "storyboard_option.update", "storyboard_option_item", item.id, before, _option_summary(item))
    await db.commit()
    return _option_summary(item)


@router.delete("/storyboard-options/{item_id}")
async def storyboard_option_delete(item_id: str, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """软删除；genre 级联软删除子孙，避免树断链孤儿。历史项目 config 存中文名不受影响。"""
    require_permission(user, STORYBOARD_OPTIONS_MANAGE)
    item = await db.get(StoryboardOptionItemModel, item_id)
    if not item or item.deleted_at is not None:
        raise HTTPException(404, "选项不存在")
    now = utcnow()
    targets = [item]
    if item.kind == "genre":
        frontier = [item.id]
        while frontier:
            children = list(
                (await db.execute(select(StoryboardOptionItemModel).where(StoryboardOptionItemModel.parent_id.in_(frontier), StoryboardOptionItemModel.deleted_at.is_(None))))
                .scalars()
                .all()
            )
            targets.extend(children)
            frontier = [child.id for child in children]
    for target in targets:
        target.deleted_at = now
    await audit(db, request, user, "storyboard_option.delete", "storyboard_option_item", item.id, _option_summary(item), {"cascadeCount": len(targets) - 1})
    await db.commit()
    return {"ok": True, "cascadeCount": len(targets) - 1}


# ---------------------------------------------------------------------------
# RunningHub 云端工作流测试（MiniMax H3 多合一）
# ---------------------------------------------------------------------------


def _runninghub_guard() -> None:
    if not settings.runninghub_api_key:
        raise HTTPException(503, "RunningHub API Key 未配置，请在 backend/.env 设置 RUNNINGHUB_API_KEY 后重启后端")


def _runninghub_error(exc: RunningHubError) -> HTTPException:
    return HTTPException(502, str(exc))


def _h3_preset_json(item: H3TestPresetModel) -> dict[str, Any]:
    return {
        "id": item.id,
        "name": item.name,
        "mode": item.mode,
        "comparisonMode": (item.usage_data or {}).get("comparisonMode", item.mode),
        "prompt": item.prompt,
        "duration": item.duration,
        "aspectRatio": item.aspect_ratio,
        "inputMedia": item.input_media,
        "outputMedia": item.output_media,
        "taskId": item.task_id,
        "taskStatus": item.task_status,
        "usage": item.usage_data,
        "createdAt": iso(item.created_at),
    }


def _h3_first_frame_ratio(ratio: str) -> str:
    return {
        "16:9": "16:9 (Widescreen)",
        "9:16": "9:16 (Portrait Widescreen)",
        "1:1": "1:1 (Square)",
        "4:3": "4:3 (Classic)",
        "3:4": "3:4 (Portrait Standard)",
    }.get(ratio, "16:9 (Widescreen)")


@router.get("/runninghub/presets")
async def runninghub_presets(user: CurrentUser, db: AsyncSession = Db):
    """读取当前管理员自己的持久化 H3 测试输入和 TOS 输出。"""
    require_admin(user)
    rows = (
        await db.execute(
            select(H3TestPresetModel)
            .where(H3TestPresetModel.user_id == user.id, H3TestPresetModel.deleted_at.is_(None))
            .order_by(H3TestPresetModel.sort_order, H3TestPresetModel.created_at.desc())
        )
    ).scalars()
    return {"items": [_h3_preset_json(item) for item in rows]}


@router.get("/runninghub/status")
async def runninghub_status(user: CurrentUser):
    """测试页初始化信息：key 是否已配置（只回显尾号）、工作流 ID、可选参数。"""
    require_admin(user)
    key = settings.runninghub_api_key
    return {
        "configured": bool(key),
        "keyTail": f"...{key[-4:]}" if len(key) >= 4 else "",
        "workflowId": settings.runninghub_workflow_id,
        "modes": ["reference", "text", "first_frame"],
        "aspectRatios": list(ASPECT_RATIOS),
        "firstFrameAspectRatios": list(FIRST_FRAME_ASPECT_RATIOS),
        "textAspectRatios": list(TEXT_ASPECT_RATIOS),
        "durationRange": [MIN_DURATION, MAX_DURATION],
        # 一/二阶段共用同一张 megapixels → 16:9 分辨率表
        "megapixelsPresets": [{"value": value, "size": size} for value, size in MEGAPIXELS_PRESETS_16X9],
        "megapixelsDefault": [DEFAULT_STAGE1_MEGAPIXELS, DEFAULT_STAGE2_MEGAPIXELS],
        "textMegapixelsDefault": DEFAULT_TEXT_MEGAPIXELS,
        "firstFrameMegapixelsDefault": DEFAULT_FIRST_FRAME_MEGAPIXELS,
    }


@router.get("/runninghub/comparison-sources")
async def runninghub_comparison_sources(user: CurrentUser, db: AsyncSession = Db):
    """列出所有用户已生成 Seedance 2.0 成片且具备公网首帧的通用分镜镜头。"""
    require_admin(user)
    rows = (
        await db.execute(
            select(UserModel, ProjectModel, ProjectTaskModel, StoryboardLineModel, ShotAssetModel, DigitalHumanModel)
            .join(ProjectModel, ProjectModel.user_id == UserModel.id)
            .join(ProjectTaskModel, ProjectTaskModel.project_id == ProjectModel.id)
            .join(StoryboardLineModel, StoryboardLineModel.project_task_id == ProjectTaskModel.id)
            .join(ShotAssetModel, ShotAssetModel.storyboard_line_id == StoryboardLineModel.id)
            .join(ProjectCastModel, ProjectCastModel.project_task_id == ProjectTaskModel.id, isouter=True)
            .join(DigitalHumanModel, DigitalHumanModel.id == ProjectCastModel.digital_human_id, isouter=True)
            .where(
                UserModel.deleted_at.is_(None),
                ProjectModel.deleted_at.is_(None),
                ProjectTaskModel.deleted_at.is_(None),
                ProjectTaskModel.storyboard_type == "general",
                StoryboardLineModel.deleted_at.is_(None),
                ShotAssetModel.deleted_at.is_(None),
                ShotAssetModel.is_current.is_(True),
                ShotAssetModel.cover_url != "",
                ShotAssetModel.video_url != "",
            )
            .order_by(UserModel.username, ProjectTaskModel.created_at.desc(), StoryboardLineModel.sort_order)
            .limit(500)
        )
    ).all()
    items = []
    seen: set[str] = set()
    for owner, project, task, line, asset, human in rows:
        if line.id in seen or (line.shot_options or {}).get("videoModel") != "doubao-seedance-2.0":
            continue
        seen.add(line.id)
        references = [
            {
                "id": f"{line.id}:seedance-cover",
                "label": "首帧",
                "url": asset.cover_url,
                "kind": "cover",
            }
        ]
        if human and human.avatar_url:
            references.append(
                {
                    "id": f"{line.id}:cast:{human.id}",
                    "label": human.name,
                    "url": human.asset_avatar_url or human.avatar_url,
                    "kind": "character",
                }
            )
        items.append(
            {
                "lineId": line.id,
                "lineOrder": line.sort_order,
                "shotType": line.shot_type or ("empty" if line.shot_prompt.strip().startswith(("无人", "空镜")) else "character"),
                "prompt": f"{line.scene_prompt.strip()}\n\n{line.shot_prompt.strip()}",
                "coverUrl": asset.cover_url,
                "seedanceUrl": asset.video_url,
                "duration": asset.duration,
                "username": owner.username,
                "userId": owner.id,
                "projectId": project.id,
                "projectName": project.name,
                "taskId": task.id,
                "taskTitle": task.title,
                "referenceCandidates": references,
            }
        )
    return {"items": items}


class RunningHubComparisonIn(BaseModel):
    line_id: str = Field(min_length=1, alias="lineId")
    reference_urls: list[str] = Field(default_factory=list, alias="referenceUrls")
    comparison_mode: str = Field(default="multi_reference", pattern="^(multi_reference|first_frame)$", alias="comparisonMode")

    model_config = {"populate_by_name": True}


@router.post("/runninghub/comparisons", status_code=201)
async def runninghub_comparison_create(
    payload: RunningHubComparisonIn,
    request: Request,
    user: CurrentUser,
    db: AsyncSession = Db,
):
    """复用现有 Seedance 镜头的提示词和参考图创建 H3 对比任务。"""
    require_admin(user)
    _runninghub_guard()
    row = (
        await db.execute(
            select(UserModel, ProjectModel, ProjectTaskModel, StoryboardLineModel, ShotAssetModel)
            .join(ProjectModel, ProjectModel.user_id == UserModel.id)
            .join(ProjectTaskModel, ProjectTaskModel.project_id == ProjectModel.id)
            .join(StoryboardLineModel, StoryboardLineModel.project_task_id == ProjectTaskModel.id)
            .join(ShotAssetModel, ShotAssetModel.storyboard_line_id == StoryboardLineModel.id)
            .where(
                StoryboardLineModel.id == payload.line_id,
                UserModel.deleted_at.is_(None),
                ProjectModel.deleted_at.is_(None),
                ProjectTaskModel.deleted_at.is_(None),
                ProjectTaskModel.storyboard_type == "general",
                StoryboardLineModel.deleted_at.is_(None),
                ShotAssetModel.deleted_at.is_(None),
                ShotAssetModel.is_current.is_(True),
            )
            .order_by(ShotAssetModel.created_at.desc())
            .limit(1)
        )
    ).one_or_none()
    if not row:
        raise HTTPException(404, "没有找到可对比的通用分镜视频")
    owner, project, task, line, asset = row
    if (line.shot_options or {}).get("videoModel") != "doubao-seedance-2.0" or not asset.cover_url or not asset.video_url:
        raise HTTPException(422, "该镜头不是具备首帧的 Seedance 2.0 成片")
    prompt = f"{line.scene_prompt.strip()}\n\n{line.shot_prompt.strip()}"
    duration = min(MAX_DURATION, max(MIN_DURATION, asset.duration))
    aspect_ratio = _h3_first_frame_ratio((line.shot_options or {}).get("ratio", "16:9"))
    references: list[dict[str, Any]] = [
        {
            "type": "image",
            "url": asset.cover_url,
            "name": "Seedance 首帧",
            "role": "comparison_cover",
        }
    ]
    cast_rows = (
        (
            await db.execute(
                select(DigitalHumanModel)
                .join(ProjectCastModel, ProjectCastModel.digital_human_id == DigitalHumanModel.id)
                .where(
                    ProjectCastModel.project_task_id == task.id,
                    ProjectCastModel.deleted_at.is_(None),
                    DigitalHumanModel.deleted_at.is_(None),
                    DigitalHumanModel.avatar_url.isnot(None),
                    DigitalHumanModel.avatar_url != "",
                )
                .order_by(ProjectCastModel.sort_order, DigitalHumanModel.created_at.desc())
            )
        )
        .scalars()
        .all()
    )
    for human in cast_rows:
        if len(references) >= 3:
            break
        references.append(
            {
                "type": "image",
                "url": human.asset_avatar_url or human.avatar_url,
                "name": human.name,
                "role": "cast_reference",
                "humanId": human.id,
                "assetAvatarUrl": human.asset_avatar_url,
                "avatarUrl": human.avatar_url,
            }
        )
    selected_urls = [url.strip() for url in payload.reference_urls if url.strip()]
    if selected_urls:
        lookup = {item["url"]: item for item in references}
        ordered: list[dict[str, Any]] = []
        for url in selected_urls:
            item = lookup.get(url)
            if item and item not in ordered:
                ordered.append(item)
        references = ordered or references[:1]
    if payload.comparison_mode == "first_frame":
        references = references[:1]
    if not references:
        raise HTTPException(422, "没有可用于对比的参考图")
    mode = "first_frame" if len(references) == 1 else "reference"
    prompt_with_refs = "\n".join(
        ["请严格参考下列图片并保持人物、场景和镜头风格一致："] + [f"<Picture {index + 1}> {item['name']}" for index, item in enumerate(references)] + [""] + [prompt]
    )
    submission_prompt = prompt if mode == "first_frame" else prompt_with_refs
    try:
        if mode == "first_frame":
            result = await rh_submit_first_frame_task(
                prompt=submission_prompt,
                duration=duration,
                aspect_ratio=aspect_ratio,
                image=references[0]["url"],
                seed=None,
                megapixels=DEFAULT_FIRST_FRAME_MEGAPIXELS,
            )
        else:
            result = await rh_submit_reference_task(
                prompt=prompt_with_refs,
                duration=duration,
                aspect_ratio=aspect_ratio,
                images=[item["url"] for item in references],
                seed=None,
                stage1_megapixels=DEFAULT_STAGE1_MEGAPIXELS,
                stage2_megapixels=DEFAULT_STAGE2_MEGAPIXELS,
            )
    except RunningHubError as exc:
        raise _runninghub_error(exc) from exc
    task_id = str(result.get("taskId"))
    preset = H3TestPresetModel(
        id=f"h3test-{uuid.uuid4().hex}",
        user_id=user.id,
        name=f"对比 · {owner.username} · {task.title} · 镜头 {line.sort_order + 1} · {mode}",
        mode=mode,
        prompt=submission_prompt,
        duration=duration,
        aspect_ratio=aspect_ratio,
        input_media=references
        + [
            {
                "type": "video",
                "url": asset.video_url,
                "name": "Seedance 2.0 原视频",
                "role": "seedance_source",
                "lineId": line.id,
                "lineOrder": line.sort_order,
                "shotType": line.shot_type or ("empty" if line.shot_prompt.strip().startswith(("无人", "空镜")) else "character"),
                "username": owner.username,
                "projectId": project.id,
                "projectName": project.name,
                "taskId": task.id,
                "taskTitle": task.title,
            },
        ],
        output_media=[],
        task_id=task_id,
        task_status=result.get("status", "QUEUED"),
        usage_data={
            "comparison": True,
            "comparisonMode": mode,
            "referenceCount": len(references),
            "referenceUrls": [item["url"] for item in references],
        },
    )
    db.add(preset)
    await audit(
        db,
        request,
        user,
        "runninghub_comparison.submit",
        "h3_test_preset",
        preset.id,
        None,
        {"lineId": line.id, "sourceUserId": owner.id, "runningHubTaskId": task_id},
    )
    await db.commit()
    return _h3_preset_json(preset)


@router.post("/runninghub/upload")
async def runninghub_upload(file: UploadFile, user: CurrentUser):
    """同一份测试媒体同时持久化到 TOS、上传 RunningHub。"""
    require_admin(user)
    _runninghub_guard()
    content = await file.read()
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(422, "文件超过 100MB 限制")
    try:
        filename = file.filename or "upload.bin"
        runninghub = await rh_upload_media(content, filename)
        tos_url = await get_storage().put_bytes(safe_key("h3-tests/inputs", filename), content, file.content_type)
        return {**runninghub, "tosUrl": tos_url}
    except RunningHubError as exc:
        raise _runninghub_error(exc) from exc
    except RuntimeError as exc:
        raise HTTPException(503, str(exc)) from exc


class RunningHubTaskIn(BaseModel):
    mode: str = Field(default="reference", pattern="^(reference|text|first_frame)$")
    prompt: str = Field(min_length=1, max_length=8000)
    duration: float = Field(default=8, ge=MIN_DURATION, le=MAX_DURATION)
    aspect_ratio: str = Field(default="16:9 (Widescreen)", alias="aspectRatio")
    images: list[str] = Field(default_factory=list, max_length=9)
    videos: list[str] = Field(default_factory=list, max_length=3)
    audios: list[str] = Field(default_factory=list, max_length=3)
    seed: int | None = Field(default=None, ge=0)
    stage1_megapixels: float = Field(default=DEFAULT_STAGE1_MEGAPIXELS, ge=MEGAPIXELS_MIN, le=MEGAPIXELS_MAX, alias="stage1Megapixels")
    stage2_megapixels: float = Field(default=DEFAULT_STAGE2_MEGAPIXELS, ge=MEGAPIXELS_MIN, le=MEGAPIXELS_MAX, alias="stage2Megapixels")
    text_megapixels: float = Field(default=DEFAULT_TEXT_MEGAPIXELS, ge=MEGAPIXELS_MIN, le=MEGAPIXELS_MAX, alias="textMegapixels")
    first_frame_megapixels: float = Field(default=DEFAULT_FIRST_FRAME_MEGAPIXELS, ge=MEGAPIXELS_MIN, le=MEGAPIXELS_MAX, alias="firstFrameMegapixels")

    model_config = {"populate_by_name": True}


@router.post("/runninghub/tasks", status_code=201)
async def runninghub_task_create(payload: RunningHubTaskIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """提交 MiniMax H3 工作流任务（nodeInfoList 覆盖提示词/时长/宽高比/参考图/种子）。"""
    require_admin(user)
    _runninghub_guard()
    try:
        if payload.mode == "first_frame":
            if not payload.images:
                raise RunningHubError("首帧生视频至少需要 1 张图片")
            result = await rh_submit_first_frame_task(
                prompt=payload.prompt,
                duration=payload.duration,
                aspect_ratio=payload.aspect_ratio,
                image=payload.images[0],
                seed=payload.seed,
                megapixels=payload.first_frame_megapixels,
            )
        elif payload.mode == "text":
            result = await rh_submit_text_task(
                prompt=payload.prompt,
                duration=payload.duration,
                aspect_ratio=payload.aspect_ratio,
                seed=payload.seed,
                megapixels=payload.text_megapixels,
            )
        else:
            result = await rh_submit_reference_task(
                prompt=payload.prompt,
                duration=payload.duration,
                aspect_ratio=payload.aspect_ratio,
                images=payload.images,
                videos=payload.videos,
                audios=payload.audios,
                seed=payload.seed,
                stage1_megapixels=payload.stage1_megapixels,
                stage2_megapixels=payload.stage2_megapixels,
            )
    except RunningHubError as exc:
        raise _runninghub_error(exc) from exc
    task_id = str(result.get("taskId"))
    await audit(
        db,
        request,
        user,
        "runninghub_task.submit",
        "runninghub_task",
        task_id,
        None,
        {
            "mode": payload.mode,
            "duration": payload.duration,
            "aspectRatio": payload.aspect_ratio,
            "imageCount": len(payload.images),
            "videoCount": len(payload.videos),
            "audioCount": len(payload.audios),
            "stage1Megapixels": payload.stage1_megapixels,
            "stage2Megapixels": payload.stage2_megapixels,
            "textMegapixels": payload.text_megapixels,
            "firstFrameMegapixels": payload.first_frame_megapixels,
        },
    )
    db.add(
        H3TestPresetModel(
            id=f"h3test-{uuid.uuid4().hex}",
            user_id=user.id,
            name={"reference": "多参考生成", "text": "纯文本生成", "first_frame": "首帧生成"}[payload.mode],
            mode=payload.mode,
            prompt=payload.prompt,
            duration=payload.duration,
            aspect_ratio=payload.aspect_ratio,
            input_media=[
                *[{"type": "image", "runningHubFileName": value, "url": value if value.startswith("https://") else ""} for value in payload.images],
                *[{"type": "video", "runningHubFileName": value, "url": value if value.startswith("https://") else ""} for value in payload.videos],
                *[{"type": "audio", "runningHubFileName": value, "url": value if value.startswith("https://") else ""} for value in payload.audios],
            ],
            output_media=[],
            task_id=task_id,
            task_status=result.get("status", "QUEUED"),
            usage_data={},
        )
    )
    await db.commit()
    return {"taskId": task_id, "status": result.get("status", "")}


class RunningHubQueryIn(BaseModel):
    task_id: str = Field(min_length=1, alias="taskId")

    model_config = {"populate_by_name": True}


@router.post("/runninghub/query")
async def runninghub_task_query(payload: RunningHubQueryIn, user: CurrentUser, db: AsyncSession = Db):
    """查询任务状态，透传 status/results/usage/errorMessage。"""
    require_admin(user)
    _runninghub_guard()
    try:
        result = await rh_query_task(payload.task_id)
    except RunningHubError as exc:
        raise _runninghub_error(exc) from exc
    preset = (
        await db.execute(
            select(H3TestPresetModel).where(
                H3TestPresetModel.user_id == user.id,
                H3TestPresetModel.task_id == payload.task_id,
                H3TestPresetModel.deleted_at.is_(None),
            )
        )
    ).scalar_one_or_none()
    if preset:
        preset.task_status = result.get("status", preset.task_status)
        preset.usage_data = result.get("usage") or {}
        if result.get("status") == "SUCCESS" and not preset.output_media:
            archived = []
            for index, output in enumerate(result.get("results") or []):
                source_url = output.get("url", "")
                if not source_url:
                    continue
                tos_url = await import_remote(source_url, f"h3-tests/videos/{user.id}", f"{payload.task_id}-{index}.mp4")
                archived.append({**output, "sourceUrl": source_url, "url": tos_url})
            preset.output_media = archived
        if preset.output_media:
            result["results"] = preset.output_media
        await db.commit()
    return result


# ---------------------------------------------------------------------------
# Kling V3 Omni 视频模型测试
# ---------------------------------------------------------------------------


def _kling_guard() -> None:
    if not settings.kling_api_key:
        raise HTTPException(503, "Kling API Key 未配置，请设置 KLING_API_KEY（或 VIDEO_API_KEY/共享 AIGC_TOKEN）后重启后端")


def _kling_error(exc: KlingError) -> HTTPException:
    return HTTPException(502, str(exc))


@router.get("/kling/status")
async def kling_status(user: CurrentUser):
    """测试页初始化信息：key 是否已配置（只回显尾号）、模型与可选参数。"""
    require_admin(user)
    key = settings.kling_api_key
    return {
        "configured": bool(key),
        "keyTail": f"...{key[-4:]}" if len(key) >= 4 else "",
        "baseUrl": settings.kling_api_base_url,
        "model": settings.kling_model,
        "modes": list(KLING_MODES),
        "aspectRatios": list(KLING_ASPECT_RATIOS),
        "imageTypes": list(KLING_IMAGE_TYPES),
        "durationRange": [KLING_MIN_DURATION, KLING_MAX_DURATION],
    }


class KlingImageIn(BaseModel):
    image_url: str = Field(min_length=1, alias="imageUrl")
    type: str = "reference"

    model_config = {"populate_by_name": True}


class KlingVideoIn(BaseModel):
    video_url: str = Field(min_length=1, alias="videoUrl")
    refer_type: str | None = Field(default=None, alias="referType")
    keep_original_sound: str | None = Field(default=None, alias="keepOriginalSound")

    model_config = {"populate_by_name": True}


class KlingTaskIn(BaseModel):
    prompt: str = Field(default="", max_length=8000)
    negative_prompt: str = Field(default="", max_length=2000, alias="negativePrompt")
    images: list[KlingImageIn] = Field(default_factory=list, max_length=4)
    videos: list[KlingVideoIn] = Field(default_factory=list, max_length=2)
    element_ids: list[str] = Field(default_factory=list, max_length=4, alias="elementIds")
    duration: float = Field(default=5, ge=KLING_MIN_DURATION, le=KLING_MAX_DURATION)
    mode: str = "pro"
    aspect_ratio: str = Field(default="16:9", alias="aspectRatio")
    sound: str = "off"
    cfg_scale: float = Field(default=0.5, ge=0, le=1, alias="cfgScale")

    model_config = {"populate_by_name": True}


@router.post("/kling/tasks", status_code=201)
async def kling_task_create(payload: KlingTaskIn, request: Request, user: CurrentUser, db: AsyncSession = Db):
    """创建 Kling V3 Omni 生成任务（文生/图生/首尾帧/多模态参考）。"""
    require_admin(user)
    _kling_guard()
    try:
        result = await kling_create_task(
            prompt=payload.prompt,
            negative_prompt=payload.negative_prompt,
            images=[item.model_dump(by_alias=True) for item in payload.images],
            videos=[item.model_dump(by_alias=True, exclude_none=True) for item in payload.videos],
            element_ids=payload.element_ids,
            duration=payload.duration,
            mode=payload.mode,
            aspect_ratio=payload.aspect_ratio,
            sound=payload.sound,
            cfg_scale=payload.cfg_scale,
        )
    except KlingError as exc:
        raise _kling_error(exc) from exc
    await audit(
        db,
        request,
        user,
        "kling_task.submit",
        "kling_task",
        result["taskId"],
        None,
        {
            "duration": payload.duration,
            "mode": payload.mode,
            "aspectRatio": payload.aspect_ratio,
            "sound": payload.sound,
            "imageCount": len(payload.images),
            "videoCount": len(payload.videos),
        },
    )
    await db.commit()
    return result


@router.get("/kling/tasks/{task_id}")
async def kling_task_query(task_id: str, user: CurrentUser):
    """查询任务状态，透传 task_status/task_result/task_status_msg。"""
    require_admin(user)
    _kling_guard()
    try:
        return await kling_query_task(task_id)
    except KlingError as exc:
        raise _kling_error(exc) from exc
